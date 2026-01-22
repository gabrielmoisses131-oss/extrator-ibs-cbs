# -*- coding: utf-8 -*-
"""
Extrator XML -> Planilha (IBS/CBS)
- Faz upload de 1+ XML (NFe/NFCe) e opcionalmente uma planilha modelo (.xlsx)
- Extrai: Data, Número da Nota, Item/Serviço, cClassTrib, Base (vBC), vIBS, vCBS, arquivo, Fonte do valor
- Grava na aba "LANCAMENTOS" preservando fórmulas/validações existentes (Excel recalcula ao abrir)

Como rodar:
  python -m pip install -r requirements.txt
  python -m streamlit run app.py
"""
import io
import zipfile
from datetime import datetime, date
import xml.etree.ElementTree as ET

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from textwrap import dedent

# -----------------------------
# Page config + CSS (Figma-like)
# -----------------------------
st.set_page_config(page_title="Extrator XML - IBS/CBS", layout="wide")

CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

:root{
  --card: rgba(255,255,255,.92);
  --card2: rgba(255,255,255,.82);
  --ink: #0f172a;
  --muted:#64748b;
  --line: rgba(15,23,42,.10);
  --shadow: 0 18px 45px rgba(2,6,23,.10);
  --shadow2: 0 26px 70px rgba(2,6,23,.16);
  --radius: 18px;

  --blue:#2563eb;
  --green:#16a34a;
  --amber:#f59e0b;
  --purple:#7c3aed;
}

.stApp{
  background:
    radial-gradient(1200px 520px at 12% -10%, rgba(37,99,235,.18), transparent 45%),
    radial-gradient(900px 520px at 110% 10%, rgba(124,58,237,.18), transparent 50%),
    radial-gradient(900px 520px at 40% 120%, rgba(22,163,74,.14), transparent 45%),
    #f6f8fc !important;
  font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif !important;
}

.block-container{
  padding-top: 1.5rem !important;
  padding-bottom: 2.5rem !important;
  max-width: 1200px !important;
}

h1,h2,h3,h4,h5,h6{ color: var(--ink) !important; letter-spacing: -.02em; }
p,li,span,small,.stCaption{ color: var(--muted) !important; }

/* Sidebar (coluna arredondada estilo app) */
section[data-testid="stSidebar"]{
  background: radial-gradient(800px 520px at 30% 0%, rgba(37,99,235,.35), transparent 50%),
              radial-gradient(800px 520px at 70% 80%, rgba(124,58,237,.35), transparent 55%),
              #0b1220 !important;

  /* Contorno + formato de “painel” */
  border: 1px solid rgba(255,255,255,.12) !important;
  border-radius: 22px !important;
  box-shadow: 0 20px 55px rgba(0,0,0,.35), inset 0 0 0 1px rgba(255,255,255,.05) !important;

  /* Respiro para parecer coluna flutuante */
  margin: 12px !important;
  overflow: hidden !important;
}

/* Garante que o conteúdo interno respeite o arredondado e ocupe a altura toda */
section[data-testid="stSidebar"] > div{
  border-radius: 22px !important;
  overflow: hidden !important;
  height: calc(100vh - 24px) !important;
}

/* Texto/cores dentro da sidebar */
section[data-testid="stSidebar"] *{ color: rgba(255,255,255,.92) !important; }
section[data-testid="stSidebar"] .stCaption,
section[data-testid="stSidebar"] small{ color: rgba(255,255,255,.65) !important; }

/* Cards */
.card{
  background: linear-gradient(180deg, var(--card), var(--card2));
  border: 1px solid var(--line);
  border-radius: var(--radius);
  box-shadow: var(--shadow);
  padding: 18px 20px;
  backdrop-filter: blur(10px);
}
.card + .card{ margin-top: 14px; }

/* Top header (match premium mock) */
.topbar{
  background: linear-gradient(180deg, rgba(255,255,255,.96), rgba(255,255,255,.86));
  border: 1px solid rgba(15,23,42,.10);
  border-radius: 18px;
  box-shadow: var(--shadow);
  padding: 16px 18px;
  display:flex;
  align-items:center;
  justify-content:space-between;
  gap: 16px;
}
.topbar .brand{ display:flex; align-items:center; gap: 12px; }
.brand-badge{
  width: 44px;
  height: 44px;
  border-radius: 16px;
  background: rgba(15,23,42,.04);
  border: 1px solid rgba(15,23,42,.08);
  display:flex;
  align-items:center;
  justify-content:center;
  flex: 0 0 auto;
}
.brand-badge svg{ width: 22px; height: 22px; }
.topbar h1{ margin:0; font-size: 1.45rem; font-weight: 900; color:#0f172a; }
.topbar .sub{ margin-top:2px; font-size:.92rem; color:#64748b; }
.status-pill{
  display:inline-flex;
  align-items:center;
  gap: 8px;
  padding: 8px 12px;
  border-radius: 999px;
  background: rgba(15,23,42,.04);
  border: 1px solid rgba(15,23,42,.08);
  font-weight: 800;
  color:#475569;
  white-space: nowrap;
}
.status-dot{ width:8px; height:8px; border-radius: 999px; background: #64748b; }

.hr{ height:1px; background: rgba(15,23,42,.10); margin: 18px 0; }

.pill{
  display:inline-flex;
  align-items:center;
  gap:8px;
  padding: 7px 11px;
  border-radius: 999px;
  background: rgba(15,23,42,.05);
  border: 1px solid rgba(15,23,42,.08);
  color: var(--muted);
  font-weight: 700;
  font-size: .82rem;
}

/* KPI grid + clickable cards */
.kpi-grid{ display:grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 16px; }
@media(max-width:1200px){ .kpi-grid{ grid-template-columns: repeat(2, 1fr);} }
@media(max-width:650px){ .kpi-grid{ grid-template-columns: 1fr;} }

.kpi-link{ text-decoration:none !important; color: inherit !important; display:block; }

.kpi{
  background: linear-gradient(180deg, rgba(255,255,255,.96), rgba(255,255,255,.86));
  border: 1px solid rgba(15,23,42,.10);
  border-radius: 18px;
  box-shadow: var(--shadow);
  padding: 16px 18px;
  position: relative;
  overflow:hidden;
  transition: transform .22s ease, box-shadow .22s ease, filter .22s ease;
  cursor: pointer;
}

.kpi::before{
  content:"";
  position:absolute; left:0; top:0; bottom:0; width: 5px;
  background: #cbd5e1;
}
.kpi.kpi-ibs::before{ background: var(--blue); }
.kpi.kpi-cbs::before{ background: var(--green); }
.kpi.kpi-cred::before{ background: var(--amber); }
.kpi.kpi-total::before{ background: var(--purple); }

.kpi::after{
  content:"";
  position:absolute;
  width: 240px; height: 240px;
  right:-90px; top:-110px;
  border-radius: 999px;
  opacity: .55;
  background: radial-gradient(circle at 30% 30%, rgba(37,99,235,.20), transparent 60%);
}
.kpi.kpi-cbs::after{ background: radial-gradient(circle at 30% 30%, rgba(22,163,74,.22), transparent 60%); }
.kpi.kpi-cred::after{ background: radial-gradient(circle at 30% 30%, rgba(245,158,11,.26), transparent 60%); }
.kpi.kpi-total::after{ background: radial-gradient(circle at 30% 30%, rgba(124,58,237,.22), transparent 60%); }

.kpi:hover{ transform: translateY(-6px); box-shadow: var(--shadow2); }
.kpi:active{ transform: translateY(-2px); box-shadow: var(--shadow); }

.kpi.is-active{
  outline: 3px solid rgba(15,23,42,.10);
  box-shadow: var(--shadow2);
  transform: translateY(-4px);
}

.kpi-head{ display:flex; align-items:flex-start; justify-content:space-between; gap: 12px; margin-bottom: 8px; position: relative; z-index: 1; }
.kpi-icon{
  width: 40px; height: 40px; border-radius: 14px;
  border: 1px solid rgba(15,23,42,.08);
  display:flex; align-items:center; justify-content:center;
  background: rgba(255,255,255,.72);
  box-shadow: 0 10px 25px rgba(2,6,23,.08);
}
.kpi-icon svg{ width: 18px; height: 18px; opacity:.95; }

.kpi .label{ color: var(--muted); font-size: .90rem; font-weight: 700; }
.kpi .value{ color: var(--ink); font-size: 1.75rem; font-weight: 900; letter-spacing: -0.02em; position: relative; z-index: 1; }
.kpi .sub{ color: var(--muted); font-size: .86rem; margin-top: 4px; position: relative; z-index: 1; }

/* Panels */
.panel-title{ display:flex; align-items:flex-start; gap: 10px; margin-bottom: 8px; }
.panel-title h3{ margin:0; font-size: 1.05rem; color: var(--ink) !important; }
.panel-title .hint{ color: var(--muted); font-size: 0.86rem; margin-top: 2px; }

.icon{
  width: 34px; height: 34px; border-radius: 12px;
  border: 1px solid rgba(15,23,42,.08);
  display:flex; align-items:center; justify-content:center;
  background: rgba(255,255,255,.78);
  box-shadow: 0 10px 25px rgba(2,6,23,.08);
}
.icon svg{ width: 18px; height: 18px; opacity:.95; }

.bar-track{ height: 10px; background: rgba(15,23,42,.06); border-radius: 999px; overflow:hidden; border: 1px solid rgba(15,23,42,.07);}
.bar-fill{ height:100%; border-radius: 999px; }
.bar-fill.ibs{ background: var(--blue); }
.bar-fill.cbs{ background: var(--green); }
.bar-fill.cred{ background: var(--amber); }

.bar-label{ display:flex; justify-content:space-between; align-items:center; font-size: 0.92rem; color: var(--muted); margin-bottom: 6px; }
.bar-foot{ display:flex; justify-content:space-between; align-items:center; margin-top: 10px; padding-top: 10px; border-top:1px solid rgba(15,23,42,.10); }
.badge-money{ font-weight: 900; }

/* Buttons */
.stButton>button, .stDownloadButton>button{
  background: linear-gradient(135deg, #111827, #0f172a) !important;
  color: #fff !important;
  border: 1px solid rgba(255,255,255,.10) !important;
  border-radius: 14px !important;
  padding: 10px 14px !important;
  font-weight: 900 !important;
  box-shadow: 0 14px 35px rgba(2,6,23,.20) !important;
  transition: transform .2s ease, box-shadow .2s ease, filter .2s ease !important;
}
.stButton>button:hover, .stDownloadButton>button:hover{
  transform: translateY(-2px) !important;
  box-shadow: 0 22px 55px rgba(2,6,23,.26) !important;
  filter: brightness(1.03) !important;
}
.stButton>button:active, .stDownloadButton>button:active{ transform: translateY(0px) !important; }

/* Inputs */
.stTextInput input, .stDateInput input{
  border-radius: 14px !important;
  border: 1px solid rgba(15,23,42,.12) !important;
  box-shadow: 0 10px 25px rgba(2,6,23,.06) !important;
}
.stSelectbox div[data-baseweb="select"] > div{
  border-radius: 14px !important;
  border: 1px solid rgba(15,23,42,.12) !important;
  box-shadow: 0 10px 25px rgba(2,6,23,.06) !important;
}

/* DataFrame */
.stDataFrame{
  border-radius: 16px !important;
  overflow:hidden !important;
  border: 1px solid rgba(15,23,42,.10) !important;
  box-shadow: 0 18px 45px rgba(2,6,23,.10) !important;
}

/* Uploader custom card */
.uploader-box{
  background: rgba(255,255,255,.06);
  border: 1px solid rgba(255,255,255,.10);
  border-radius: 18px;
  padding: 16px;
  box-shadow: 0 18px 40px rgba(0,0,0,.25);
}

/* === FIX: remove decorative giant icons === */
.kpi::after{ display: none !important; }

/* Tip (Dica importante) – premium + icon sized correctly */
.tip{
  display:flex;
  gap: 12px;
  align-items:flex-start;
  padding: 14px 16px;
  border-radius: 16px;
  background: #fff7ed;
  border: 1px solid rgba(180,83,9,.18);
  box-shadow: 0 12px 35px rgba(2,6,23,.06);
}
.tip strong{ display:block; color:#b45309; font-weight:900; margin-bottom:2px; }
.tip span{ color:#92400e !important; font-size:.92rem; }
.tip-icon{
  width: 36px;
  height: 36px;
  border-radius: 14px;
  background: rgba(245,158,11,.18);
  border: 1px solid rgba(245,158,11,.22);
  display:flex;
  align-items:center;
  justify-content:center;
  flex: 0 0 auto;
}
.tip-icon svg { width: 18px; height: 18px; }

/* ===== FIX UPLOAD ZONA BRANCA (SIDEBAR) ===== */
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"]{
  background: rgba(255,255,255,.06) !important;
  border: 1px dashed rgba(255,255,255,.22) !important;
  border-radius: 18px !important;
  padding: 14px !important;
}

section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] *{
  color: rgba(255,255,255,.90) !important;
}

section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] svg{
  fill: rgba(255,255,255,.90) !important;
  color: rgba(255,255,255,.90) !important;
  opacity: 1 !important;
}

section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] button{
  background: rgba(255,255,255,.10) !important;
  border: 1px solid rgba(255,255,255,.18) !important;
  color: rgba(255,255,255,.92) !important;
  border-radius: 12px !important;
}

/* ===== TABELA PREMIUM (igual vídeo) ===== */
.table-wrap{
  background: rgba(255,255,255,.92);
  border: 1px solid rgba(15,23,42,.10);
  border-radius: 18px;
  box-shadow: 0 18px 45px rgba(2,6,23,.10);
  padding: 16px;
  margin-top: 10px;
}

</style>
"""

st.markdown(CSS, unsafe_allow_html=True)

# -----------------------------
# XML helpers
# -----------------------------
def _local(tag: str) -> str:
    # "{ns}Tag" -> "Tag"
    return tag.split("}", 1)[-1] if "}" in tag else tag

def _find_text(elem: ET.Element, path: str) -> str | None:
    x = elem.find(path)
    if x is None or x.text is None:
        return None
    return x.text.strip()

def _parse_date(root: ET.Element) -> date | None:
    """
    Tenta pegar data de emissão:
      - NFe/infNFe/ide/dhEmi (ISO datetime) ou dEmi (YYYY-MM-DD)
    """
    for p in [
        ".//{*}infNFe/{*}ide/{*}dhEmi",
        ".//{*}infNFe/{*}ide/{*}dEmi",
        ".//{*}ide/{*}dhEmi",
        ".//{*}ide/{*}dEmi",
    ]:
        t = _find_text(root, p)
        if not t:
            continue
        try:
            # dhEmi pode ser "2026-01-08T10:22:33-03:00"
            if "T" in t:
                # remove timezone para parse mais simples
                base = t.split("T")[0]
                return datetime.fromisoformat(base).date() if len(base) > 10 else datetime.fromisoformat(t[:19]).date()
            return datetime.fromisoformat(t).date()
        except Exception:
            try:
                return datetime.strptime(t[:10], "%Y-%m-%d").date()
            except Exception:
                pass
    return None

def _parse_nnf(root: ET.Element) -> str | None:
    # Número da NF: ide/nNF
    for p in [".//{*}infNFe/{*}ide/{*}nNF", ".//{*}ide/{*}nNF"]:
        t = _find_text(root, p)
        if t:
            return t
    return None

def _parse_items_from_xml(xml_bytes: bytes, filename: str) -> list[dict]:
    """
    Extrai itens (det) e IBS/CBS:
      - Item/Serviço: det/prod/xProd
      - cClassTrib: imposto/IBSCBS/cClassTrib
      - Base (vBC): imposto/IBSCBS/vBC
      - vIBS / vCBS: imposto/IBSCBS/vIBS, vCBS (se existirem)
    """
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return []

    emissao = _parse_date(root)
    nnf = _parse_nnf(root)

    rows: list[dict] = []
    dets = root.findall(".//{*}infNFe/{*}det") or root.findall(".//{*}det")
    for det in dets:
        xprod = _find_text(det, ".//{*}prod/{*}xProd") or ""
        ibscbs = det.find(".//{*}imposto/{*}IBSCBS")
        if ibscbs is None:
            # alguns XML podem não ter IBSCBS -> ignora item
            continue

        cclass = _find_text(ibscbs, ".//{*}cClassTrib") or ""
        vbc = _find_text(ibscbs, ".//{*}vBC")
        vibs = _find_text(ibscbs, ".//{*}vIBS")
        vcbs = _find_text(ibscbs, ".//{*}vCBS")

        def _to_float(x: str | None):
            try:
                return float(x) if x not in (None, "") else None
            except Exception:
                return None

        vbc_f = _to_float(vbc)
        vibs_f = _to_float(vibs)
        vcbs_f = _to_float(vcbs)

        # Fonte do valor (base)
        fonte = "IBSCBS/vBC" if vbc_f is not None else ""

        rows.append(
            {
                "Data": emissao,
                "Numero": nnf,
                "Item/Serviço": xprod,
                "cClassTrib": cclass,
                "Valor da operação": vbc_f,
                "vIBS": vibs_f,
                "vCBS": vcbs_f,
                "arquivo": filename,
                "Fonte do valor": fonte,
            }
        )

    return rows

# -----------------------------
# Excel write helper
# -----------------------------
def _append_to_workbook(template_bytes: bytes, df: pd.DataFrame) -> bytes:
    """
    Abre o template e grava df na aba LANCAMENTOS, acrescentando linhas.

    ✅ O que este writer garante:
      - Encontra a linha correta de cabeçalhos mesmo que o layout mude (ex.: cabeçalho na linha 2).
      - Escreve nos campos de entrada (Data, Numero, Item/Serviço, etc.).
      - COPIA fórmulas/estilos da primeira linha-modelo de dados para todas as novas linhas,
        para que "Base", "Valor IBS/CBS", validações e cálculos voltem a aparecer no Excel.
    """
    from copy import copy
    bio = io.BytesIO(template_bytes)
    wb = load_workbook(bio)

    ws = wb["LANCAMENTOS"] if "LANCAMENTOS" in wb.sheetnames else wb.active

    # ------------------------------------------------------------
    # 1) Descobre em qual linha estão os cabeçalhos (layout pode mudar)
    # ------------------------------------------------------------
    expected = {"Data", "Numero", "Item/Serviço", "cClassTrib", "Valor da operação"}
    header_row = None

    # procura nos primeiros 25 rows (suficiente pro seu layout)
    for r in range(1, 26):
        values = []
        for c in range(1, 101):  # lê até 100 colunas (bem além do necessário)
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                values.append(v.strip())
        hit = len(expected.intersection(values))
        if hit >= 3:  # achou linha com a maioria dos cabeçalhos
            header_row = r
            break

    if header_row is None:
        # fallback antigo (assume linha 1)
        header_row = 1

    # mapeia "nome do cabeçalho" -> coluna
    headers: dict[str, int] = {}
    last_col = 0
    for col in range(1, 201):  # até 200 colunas
        v = ws.cell(row=header_row, column=col).value
        if isinstance(v, str) and v.strip():
            headers[v.strip()] = col
            last_col = max(last_col, col)

    # se ainda não achou nada (planilha muito custom), tenta usar as colunas usadas do sheet
    if last_col == 0:
        last_col = min(ws.max_column, 200)

    # ------------------------------------------------------------
    # 2) Define a "linha modelo" (a primeira linha de dados com fórmulas)
    #    No seu modelo: header_row=2, a linha 3 é seção, a 4 é a linha modelo.
    # ------------------------------------------------------------
    template_row = header_row + 2

    # ------------------------------------------------------------
    # 3) Descobre a próxima linha vazia olhando a coluna "Data"
    # ------------------------------------------------------------
    next_row = ws.max_row + 1
    if "Data" in headers:
        c = headers["Data"]
        r = ws.max_row
        while r >= (template_row) and ws.cell(row=r, column=c).value in (None, ""):
            r -= 1
        next_row = max(r + 1, template_row)

    # ------------------------------------------------------------
    # 4) Função para copiar estilo + fórmulas da linha modelo
    # ------------------------------------------------------------
    def _copy_row_style_and_formulas(src_row: int, dst_row: int):
        from copy import copy
        from openpyxl.formula.translate import Translator

        for col in range(1, last_col + 1):
            src = ws.cell(row=src_row, column=col)
            dst = ws.cell(row=dst_row, column=col)

            # estilos
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)

            # valor / fórmula
            if isinstance(src.value, str) and src.value.startswith("="):
                # traduz a referência da linha-modelo -> linha destino (ex.: G4 vira G7)
                try:
                    dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
                except Exception:
                    dst.value = src.value
            else:
                dst.value = src.value


    # ------------------------------------------------------------
    # 5) Escreve as linhas: primeiro replica modelo, depois grava os valores de entrada
    # ------------------------------------------------------------
    fields = [
        "Data", "Numero", "Item/Serviço", "cClassTrib",
        "Valor da operação", "vIBS", "vCBS", "arquivo", "Fonte do valor"
    ]

    for _, row in df.iterrows():
        # replica a linha modelo (fórmulas + visual)
        _copy_row_style_and_formulas(template_row, next_row)

        # agora sobrescreve somente os campos de ENTRADA
        for f in fields:
            if f not in headers:
                continue
            col = headers[f]
            val = row.get(f, None)

            cell = ws.cell(row=next_row, column=col)

            # datas
            if f == "Data" and pd.notna(val) and isinstance(val, date):
                cell.value = val
                cell.number_format = "dd/mm/yyyy"
            else:
                if pd.isna(val):
                    val = None
                cell.value = val

        next_row += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# -----------------------------
# UI
# -----------------------------
st.markdown(dedent("""
<div class="topbar">
  <div class="brand">
<div class="brand-badge" aria-hidden="true">
      <svg viewBox="0 0 24 24" fill="none">
        <path d="M7 3h7l3 3v15a2 2 0 0 1-2 2H7a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2z" stroke="#1f2937" stroke-width="1.7"/>
        <path d="M14 3v4a1 1 0 0 0 1 1h4" stroke="#1f2937" stroke-width="1.7"/>
        <path d="M8 12h8M8 16h8" stroke="#1f2937" stroke-width="1.7" stroke-linecap="round"/>
      </svg>
</div>
<div>
      <h1>Extrator XML - IBS/CBS</h1>
<div class="sub">Visualização de dados fiscais da reforma tributária</div>
</div>
  </div>

  <div class="status-pill">
<span class="status-dot"></span>
    Pronto para análise
  </div>
</div>
"""), unsafe_allow_html=True)

# Sidebar: uploads
with st.sidebar:
    st.markdown(dedent("""
<div style="padding: 10px 6px 6px 6px;">
  <div style="display:flex; align-items:center; gap:10px; margin-bottom: 10px;">
<div style="width:42px;height:42px;border-radius:14px;background:rgba(108,124,255,.18);border:1px solid rgba(255,255,255,.12);display:flex;align-items:center;justify-content:center;">
<span style="font-weight:900;">✦</span>
</div>
<div>
<div style="font-weight:900; font-size: 1.05rem; line-height:1;">Extrator XML</div>
<div style="font-size:.82rem; color: rgba(226,232,240,.75); margin-top:2px;">IBS/CBS</div>
</div>
<div style="margin-left:auto; font-size:.75rem; font-weight:800; padding:4px 10px; border-radius:999px; background:rgba(255,255,255,.10); border:1px solid rgba(255,255,255,.10); color: rgba(226,232,240,.9);">v2.0</div>
  </div>

  <div class="sidebar-card">
<div class="sidebar-title">
<h4>EXCEL IBS/CBS</h4>
<span class="tag">FIXA</span>
</div>

<div class="uploader-box">
<div style="text-align:center; font-weight:800; margin-bottom: 6px;">Planilha interna</div>
<div style="text-align:center; color: rgba(226,232,240,.75); font-size:.82rem;">O app usa <b>planilha_modelo.xlsx</b></div>
</div>

<div style="height: 14px;"></div>

<div class="sidebar-title" style="margin-top: 2px;">

<h4>ARQUIVOS XML</h4>
<span class="tag" style="background:rgba(37,99,235,.18); border-color: rgba(37,99,235,.22);">OBRIGATÓRIO</span>
</div>

<div class="uploader-box">
<div style="text-align:center; font-weight:800; margin-bottom: 6px;">Arraste e solte</div>
<div style="text-align:center; color: rgba(226,232,240,.75); font-size:.82rem;">ou clique para selecionar</div>
<div style="margin-top:10px;"></div>
"""), unsafe_allow_html=True)
    xml_files = st.file_uploader("XML(s)", type=["xml", "zip"], accept_multiple_files=True, label_visibility="collapsed")
    st.markdown(dedent("""
<div class="uploader-help">XML, ZIP • Múltiplos</div>
</div>

<div style="height: 14px;"></div>

<div style="padding: 12px; border-radius: 16px; background: rgba(255,255,255,.05); border: 1px solid rgba(255,255,255,.10);">
<div style="font-weight:900; margin-bottom: 4px;">Dica rápida</div>
<div style="font-size: .82rem; color: rgba(226,232,240,.78);">
        Envie XMLs para extrair automaticamente dados de IBS e CBS da reforma tributária.
</div>
</div>
  </div>
</div>
"""), unsafe_allow_html=True)

# Carrega planilha modelo FIXA (arquivo na pasta do projeto)
from pathlib import Path
TEMPLATE_PATH = Path(__file__).parent / "planilha_modelo.xlsx"

try:
    template_bytes = TEMPLATE_PATH.read_bytes()
except FileNotFoundError:
    template_bytes = None

# Aviso amigável caso o arquivo não exista (em produção ele deve estar junto do app)
if template_bytes is None:
    st.markdown(dedent("""
<div class="tip">
  <div class="tip-icon" aria-hidden="true">
    <svg viewBox="0 0 24 24" fill="none">
      <path d="M12 8v5" stroke="#b45309" stroke-width="1.8" stroke-linecap="round"/>
      <path d="M12 16h.01" stroke="#b45309" stroke-width="2.8" stroke-linecap="round"/>
      <path d="M10.3 3.7a2 2 0 0 1 3.4 0l8.4 14.7A2 2 0 0 1 20.4 21H3.6a2 2 0 0 1-1.7-3.0l8.4-14.3z"
            stroke="#b45309" stroke-width="1.6" fill="#fff7ed"/>
    </svg>
  </div>
  <div>
    <div class="tip-title">Planilha modelo não encontrada</div>
    <div class="tip-text">Coloque o arquivo <b>planilha_modelo.xlsx</b> na mesma pasta do <b>app.py</b>.</div>
  </div>
</div>
"""), unsafe_allow_html=True)

# Parse XMLs
rows_all: list[dict] = []
errors: list[str] = []
if xml_files:
    for f in xml_files:
        try:
            b = f.read()
            if f.name.lower().endswith(".zip"):
                with zipfile.ZipFile(io.BytesIO(b)) as z:
                    xml_names = [n for n in z.namelist() if n.lower().endswith(".xml")]
                    if not xml_names:
                        errors.append(f"{f.name}: zip sem .xml")
                        continue
                    for xn in xml_names:
                        xb = z.read(xn)
                        rows = _parse_items_from_xml(xb, f"{f.name}:{xn}")
                        if not rows:
                            errors.append(f"{f.name}:{xn}: não encontrei itens com IBSCBS")
                        rows_all.extend(rows)
            else:
                rows = _parse_items_from_xml(b, f.name)
                if not rows:
                    errors.append(f"{f.name}: não encontrei itens com IBSCBS")
                rows_all.extend(rows)
        except Exception as e:
            errors.append(f"{f.name}: erro ao ler ({e})")

df = pd.DataFrame(rows_all)

# Normaliza Data
if not df.empty:
    df["Data"] = pd.to_datetime(df["Data"], errors="coerce").dt.date

# ---------- KPIs ----------
def money(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return "R$ 0,00"
    try:
        return "R$ {:,.2f}".format(float(x)).replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"

def pct(x):
    try:
        return f"{float(x):.1f}%"
    except Exception:
        return ""


# --- Totais (Somatório das bases do XML) ---
# Aqui os painéis mostram apenas a SOMA DAS BASES encontradas no XML (sem aplicar alíquota).
# As alíquotas exibidas são apenas informativas (fictícias), como você pediu.
ALIQUOTA_IBS_TEXTO = "0,10%"
ALIQUOTA_CBS_TEXTO = "0,90%"

base_ibs = float(df["Valor da operação"].fillna(0).sum()) if (not df.empty and "Valor da operação" in df.columns) else 0.0
base_cbs = float(df["Valor da operação"].fillna(0).sum()) if (not df.empty and "Valor da operação" in df.columns) else 0.0

# Totais exibidos nos cards = soma das bases
ibs_total = round(base_ibs, 2)
cbs_total = round(base_cbs, 2)
total_tributos = round(base_ibs + base_cbs, 2)

# Créditos: 1% sobre UMA base (IBS ou CBS)
creditos_total = round(base_ibs * 0.01, 2)


# --- KPI clique (filtro via query param) ---
try:
    _qp = st.query_params.get("kpi", "all")
    # Streamlit pode devolver lista/tuple dependendo da versão
    if isinstance(_qp, (list, tuple)):
        selected_kpi = _qp[0] if _qp else "all"
    else:
        selected_kpi = _qp or "all"
except Exception:
    selected_kpi = "all"

selected_kpi = str(selected_kpi).lower().strip()
if selected_kpi not in ("all", "ibs", "cbs", "cred", "total"):
    selected_kpi = "all"


st.markdown(
    f"""
<div class="kpi-grid">
  <a class="kpi-link" href="?kpi=ibs">
    <div class="kpi kpi-ibs {'is-active' if selected_kpi=='ibs' else ''}">
      <div class="kpi-head">
        <div>
          <div class="label">IBS Total</div>
          <div class="pill">↗ Alíquota {ALIQUOTA_IBS_TEXTO}</div>
        </div>
        <div class="kpi-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" fill="none">
            <path d="M7 17V7" stroke="#2563eb" stroke-width="2" stroke-linecap="round"/>
            <path d="M7 17h10" stroke="#2563eb" stroke-width="2" stroke-linecap="round"/>
            <path d="M9 13l3-3 3 2 2-3" stroke="#2563eb" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"/>
          </svg>
        </div>
      </div>
      <div class="value">{money(ibs_total)}</div>
      <div class="sub">Soma das bases IBS (XML)</div>
    </div>
  </a>

  <a class="kpi-link" href="?kpi=cbs">
    <div class="kpi kpi-cbs {'is-active' if selected_kpi=='cbs' else ''}">
      <div class="kpi-head">
        <div>
          <div class="label">CBS Total</div>
          <div class="pill">↗ Alíquota {ALIQUOTA_CBS_TEXTO}</div>
        </div>
        <div class="kpi-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" fill="none">
            <path d="M8 7h8M8 12h8M8 17h8" stroke="#16a34a" stroke-width="2" stroke-linecap="round"/>
            <path d="M6 5h12a2 2 0 0 1 2 2v10a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V7a2 2 0 0 1 2-2z" stroke="#16a34a" stroke-width="2"/>
          </svg>
        </div>
      </div>
      <div class="value">{money(cbs_total)}</div>
      <div class="sub">Soma das bases CBS (XML)</div>
    </div>
  </a>

  <a class="kpi-link" href="?kpi=cred">
    <div class="kpi kpi-cred {'is-active' if selected_kpi=='cred' else ''}">
      <div class="kpi-head">
        <div>
          <div class="label">Créditos</div>
          <div class="pill">↗ IBS + CBS</div>
        </div>
        <div class="kpi-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" fill="none">
            <path d="M7 12h10" stroke="#f59e0b" stroke-width="2" stroke-linecap="round"/>
            <path d="M12 7v10" stroke="#f59e0b" stroke-width="2" stroke-linecap="round"/>
            <path d="M6 6h12v12H6z" stroke="#f59e0b" stroke-width="2" opacity=".6"/>
          </svg>
        </div>
      </div>
      <div class="value">{money(creditos_total)}</div>
      <div class="sub">Somatório de vIBS + vCBS</div>
    </div>
  </a>

  <a class="kpi-link" href="?kpi=total">
    <div class="kpi kpi-total {'is-active' if selected_kpi=='total' else ''}">
      <div class="kpi-head">
        <div>
          <div class="label">Total Tributos</div>
          <div class="pill">↗ Consolidado</div>
        </div>
        <div class="kpi-icon" aria-hidden="true">
          <svg viewBox="0 0 24 24" fill="none">
            <path d="M7 8h10M7 12h10M7 16h10" stroke="#a855f7" stroke-width="2" stroke-linecap="round"/>
            <path d="M9 3h6v3H9z" stroke="#a855f7" stroke-width="2"/>
            <path d="M6 6h12v15H6z" stroke="#a855f7" stroke-width="2" opacity=".6"/>
          </svg>
        </div>
      </div>
      <div class="value">{money(total_tributos)}</div>
      <div class="sub">IBS base + CBS base</div>
    </div>
  </a>
</div>

<div style="margin-top:10px;">
  <a class="kpi-link" href="?kpi=all"><span class="pill">Limpar filtro</span></a>
</div>
""",
    unsafe_allow_html=True,
)
# Painéis (estilo Figma) — Débitos vs Créditos
c1, c2 = st.columns(2, gap="large")
ibs_deb = float(ibs_total or 0.0)
cbs_deb = float(cbs_total or 0.0)
ibs_cred = 0.0
cbs_cred = 0.0

def _bar_width(val, vmax):
    if vmax <= 0:
        return "0%"
    return f"{max(0.0, min(1.0, val / vmax)) * 100:.1f}%"

max_ibs = max(ibs_deb, ibs_cred, 1e-9)
max_cbs = max(cbs_deb, cbs_cred, 1e-9)

with c1:
    st.markdown(
        f"""
<div class="card">
  <div class="panel-title">
<div class="panel-left">
<div class="icon" aria-hidden="true">
      <svg viewBox="0 0 24 24" fill="none">
        <path d="M4 18V6" stroke="#334155" stroke-width="1.7" stroke-linecap="round"/>
        <path d="M4 18h16" stroke="#334155" stroke-width="1.7" stroke-linecap="round"/>
        <path d="M8 14l3-3 3 2 4-5" stroke="#2563eb" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round"/>
      </svg>
</div>
<div>
      <h3>IBS - Débitos vs Créditos</h3>
<div class="hint">Imposto sobre Bens e Serviços (Estados/Municípios)</div>
</div>
</div>
<span class="badge on">Ativo</span>
  </div>

  <div class="bar-row">
<div class="bar-label"><span>Débitos</span><span class="badge-money">{money(ibs_deb)}</span></div>
<div class="bar-track"><div class="bar-fill ibs" style="width:{_bar_width(ibs_deb, max_ibs)}"></div></div>
  </div>

  <div class="bar-row">
<div class="bar-label"><span>Créditos</span><span class="badge-money">-{money(ibs_cred)}</span></div>
<div class="bar-track"><div class="bar-fill cred" style="width:{_bar_width(ibs_cred, max_ibs)}"></div></div>
  </div>

  <div class="bar-foot green">
    <strong>Saldo a Recolher</strong>
<span class="badge-money" style="color:#2563eb;">{money(ibs_deb - ibs_cred)}</span>
  </div>
</div>
""",
        unsafe_allow_html=True,
    )

with c2:
    st.markdown(
        f"""
<div class="card">
  <div class="panel-title">
<div class="panel-left">
<div class="icon" aria-hidden="true">
      <svg viewBox="0 0 24 24" fill="none">
        <path d="M4 18V6" stroke="#334155" stroke-width="1.7" stroke-linecap="round"/>
        <path d="M4 18h16" stroke="#334155" stroke-width="1.7" stroke-linecap="round"/>
        <path d="M8 14l3-3 3 2 4-5" stroke="#16a34a" stroke-width="1.9" stroke-linecap="round" stroke-linejoin="round"/>
      </svg>
</div>
<div>
      <h3>CBS - Débitos vs Créditos</h3>
<div class="hint">Contribuição sobre Bens e Serviços (União)</div>
</div>
</div>
<span class="badge on" style="background:#ecfdf3;border-color:#dcfce7;color:#166534;">Ativo</span>
  </div>

  <div class="bar-row">
<div class="bar-label"><span>Débitos</span><span class="badge-money">{money(cbs_deb)}</span></div>
<div class="bar-track"><div class="bar-fill cbs" style="width:{_bar_width(cbs_deb, max_cbs)}"></div></div>
  </div>

  <div class="bar-row">
<div class="bar-label"><span>Créditos</span><span class="badge-money">-{money(cbs_cred)}</span></div>
<div class="bar-track"><div class="bar-fill cred" style="width:{_bar_width(cbs_cred, max_cbs)}"></div></div>
  </div>

  <div class="bar-foot">
    <strong>Saldo a Recolher</strong>
<span class="badge-money" style="color:#16a34a;">{money(cbs_deb - cbs_cred)}</span>
  </div>
</div>
""",
        unsafe_allow_html=True,
    )


st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

# Alerts
if errors:
    st.warning("Alguns arquivos tiveram problemas:")
    for e in errors[:10]:
        st.write("•", e)
    if len(errors) > 10:
        st.caption(f"... e mais {len(errors)-10} itens")

# ---------- Filters + table ----------
st.markdown('<div class="card">', unsafe_allow_html=True)
st.markdown("## Itens do Documento")
st.caption("Detalhamento dos itens extraídos do XML (inclui base vBC e valores de IBS/CBS quando presentes).")

if df.empty:
    st.info("Envie XML(s) para visualizar os itens aqui.")
    st.markdown("</div>", unsafe_allow_html=True)
    st.stop()

c1, c2, c3 = st.columns([1, 2, 1], gap="large")

with c1:
    min_d = df["Data"].min()
    max_d = df["Data"].max()
    # SEMPRE define "periodo" (evita NameError)
    periodo = st.date_input("Período", value=(min_d, max_d), min_value=min_d, max_value=max_d)

with c2:
    q = st.text_input("Buscar item", placeholder="Ex.: produto, serviço, descrição...")

with c3:
    classes = sorted([c for c in df["cClassTrib"].dropna().unique().tolist() if str(c).strip() != ""])
    pick = st.selectbox("cClassTrib", options=["(Todos)"] + classes, index=0)

df_view = df.copy()

# filtro de período (robusto)
if isinstance(periodo, (list, tuple)) and len(periodo) == 2:
    d1, d2 = periodo
    df_view["Data"] = pd.to_datetime(df_view["Data"], errors="coerce").dt.date
    df_view = df_view[(df_view["Data"] >= d1) & (df_view["Data"] <= d2)]

# busca
if q:
    qq = q.strip().lower()
    df_view = df_view[df_view["Item/Serviço"].fillna("").str.lower().str.contains(qq, na=False)]

# cClassTrib
if pick and pick != "(Todos)":
    df_view = df_view[df_view["cClassTrib"].astype(str) == str(pick)]


# filtro por KPI (clique nos cards)
if selected_kpi != "all":
    vibs = df_view["vIBS"].fillna(0) if "vIBS" in df_view.columns else None
    vcbs = df_view["vCBS"].fillna(0) if "vCBS" in df_view.columns else None

    if selected_kpi == "ibs" and vibs is not None:
        df_view = df_view[vibs != 0]
    elif selected_kpi == "cbs" and vcbs is not None:
        df_view = df_view[vcbs != 0]
    elif selected_kpi == "cred" and (vibs is not None and vcbs is not None):
        # créditos normalmente aparecem como valores negativos
        df_view = df_view[(vibs < 0) | (vcbs < 0)]
    elif selected_kpi == "total" and (vibs is not None and vcbs is not None):
        df_view = df_view[(vibs != 0) | (vcbs != 0)]

show_cols = ["Data", "Numero", "Item/Serviço", "cClassTrib", "Valor da operação", "vIBS", "vCBS", "arquivo", "Fonte do valor"]
show_cols = [c for c in show_cols if c in df_view.columns]

# ===== TABELA PREMIUM (igual vídeo) =====
st.markdown('<div class="table-wrap">', unsafe_allow_html=True)

st.dataframe(
    df_view[show_cols],
    use_container_width=True,
    hide_index=True,
    height=420
)

st.download_button(
    "Baixar CSV filtrado",
    data=df_view[show_cols].to_csv(index=False).encode("utf-8"),
    file_name="itens_filtrados.csv",
    mime="text/csv",
)

st.markdown('</div>', unsafe_allow_html=True)

# ---------- Generate planilha ----------
st.markdown('<div class="hr"></div>', unsafe_allow_html=True)
st.markdown("## Gerar planilha preenchida")

if template_bytes is None:
    st.error("Não encontrei **planilha_modelo.xlsx** na mesma pasta do app.py.")
else:
    if st.button("Gerar planilha", type="primary"):
        out_bytes = _append_to_workbook(template_bytes, df_view)

        st.success("Planilha gerada! Abra no Excel para ver as fórmulas calculando.")

        st.download_button(
            "Baixar planilha_preenchida.xlsx",
            data=out_bytes,
            file_name="planilha_preenchida.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

